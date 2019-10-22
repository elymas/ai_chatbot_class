# -*- coding: utf-8 -*-
"""

"""


from openpyxl import load_workbook

def got_drink(item_dict, item):
    """
    item_dict 는 { 상품명 : 수량 } 형식으로 되어있는 Dict
    item_dict[item] ==> 해당 item에 대한 수량을 return
    Ex) print(item_dict['사이다']) ==> 10 (사이다 수량)

    item_dict[item] -= 1 은 item_dict[item] = item_dict[item] - 1 을 줄인말
    자기 자신에서 1뺀 값으로 갱신
    """
    item_dict[item] -= 1
    return 0


EXCEL_FILE_NAME = 'Database.xlsx'
db = load_workbook(filename=EXCEL_FILE_NAME)
machine_db = db['자판기']

# item_dict 라는 비어있는 딕셔너리를 선언
item_dict = {}

# machine_db의 row 를 하나씩 순차적으로 접근
# for [변수] in machine_db.rows:
# [변수]는 machine_db의 row를 하나씩 받아서 끝까지 전달
# [변수] 이름이 row여서 혼동의 소지가 있어서 설명 추가했습니다.
for row in machine_db.rows:
    # 첫번째 row의 정보는 None, 수량, 단가, 매진 / 두번째 row의 정보는 사이다, 10, 1000, None / ....
    # row의 0번째 index 값이 None 이 아니면 ==> 첫번째 열이 비어있지 않으면
    if row[0].value is not None:
        # item_dict의 [row[0].value] (사이다, 콜라, ....)
        # item_dict의 [row[1].value] ( 10, 13, ....)
        # 딕셔너리 추가하는 방법
        # 딕셔너리 이름[Key] = Value
        # 해당 딕셔너리에 {Key : Value} 가 추가
        item_dict[row[0].value] = row[1].value
        
got_drink(item_dict, '사이다')
got_drink(item_dict, '콜라')
print(item_dict['콜라'])

"""
item_dict를 다른 Excel(Database1.xlsx) 파일에 저장하는 Code 만들기
"""

EXCEL2_FILE_NAME = 'Database1.xlsx'
db2 = load_workbook(filename=EXCEL2_FILE_NAME)
machine_db2 = db2['자판기']

for row in machine_db2.rows:
    if row[0].value is not None:
        print(row[0].value)
        print(item_dict[row[0].value])
        row[1].value = item_dict[row[0].value]
        
db2.save(EXCEL2_FILE_NAME)
print('ok')