# -*- coding: utf-8 -*-
"""
Openpyxl 을 사용해서 엑셀파일에 데이터를 읽고 쓰기 

@author: Junho
"""

# openpyxl 라이브러리에서 load_workbook 을 사용하겠다고 선언
from openpyxl import load_workbook

# 사용할 엑셀 파일의 경로명 (같은 폴더안에 있어서 파일이름만 명시)
EXCEL_FILE_NAME = 'Database.xlsx'

# db라는 변수에 load_workbook 이라는 함수를 통해 엑셀 파일을 넣어준다
db = load_workbook(filename=EXCEL_FILE_NAME)

# tuto_db 라는 변수에 엑셀 파일에서 'Tuto' Sheet 를 할당 한다
tuto_db = db['Tuto']

# 엑셀에 나열된 Sheet 순서로도 접근 가능하지만 순서가 변하면 바꿔야함
# tuto_db = db.worksheets[0]


def write_with_index():
    """
    엑셀의 index 를 통해서 셀 접근 방법
    .value 를 넣어줘야 해당 셀의 값을 확인 가능
    .value 가 없으면 해당 셀의 위치정보 확인
    """
    tuto_db['A1'].value = 100
    #print(tuto_db['A1'])
    print(tuto_db['A1'].value)


def write_with_cell():
    """
    엑셀의 Cell 정보를 통해서 접근하는 방법
    .cell의 매개변수로 row, column 값을 (1~ ) 전달
    값을 수정하려면 value 매개변수로 인자를 전달
    값을 읽을때 마찬가지로 .value 를 해서 값을 읽기
    """
    tuto_db.cell(row=1, column=1, value='SBA 점심타임')
    print(tuto_db.cell(row=1, column=1))
    print(tuto_db.cell(row=1, column=1).value)


write_with_index()
write_with_cell()

# 엑셀파일을 저장
# 경로를 수정하면 다른이름으로 저장 가능 없는 파일일 경우 파일 생성
db.save(EXCEL_FILE_NAME)
print('ok')