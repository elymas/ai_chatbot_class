# -*- coding: utf-8 -*-
"""
새로운 Sheet 를 추가 

@author: Junho
"""
from openpyxl import load_workbook

EXCEL_FILE_NAME = 'Database.xlsx'
db = load_workbook(filename=EXCEL_FILE_NAME)
tuto_db = db['Tuto']

# create_sheet 로 새로운 Sheet 생성 가능
# tuto_db 는 Sheet 객체이므로, 엑셀 객체인 db를 가지고 사용
# .title 로 Sheet의 이름을 변경 가능
 create_db = db.create_sheet('생성된 시트')
 create_db.title = '바꾼 이름'

 remove_db = db.remove_sheet(db['바꾼 이름'])

db.save(EXCEL_FILE_NAME)
print('ok')